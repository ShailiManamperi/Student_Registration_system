from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib



background = "#06283D"
framebg="#EDEDED"
framefg="#06283D"

root = Tk()
root.title("Student Registration System")
root.geometry("1200x600+50+50")
root.config(bg=background)

#Creating data saving file in excel

file = pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet = file.active
    sheet['A1']="Registration No."
    sheet['B1']="Name"
    sheet['C1']="Class"
    sheet['D1']="Gender"
    sheet['E1']="DOB"
    sheet['F1']="Date of registration"
    sheet['G1']="Religion"
    sheet['H1']="Skills"
    sheet['I1']="Father's Name"
    sheet['J1']="Mother's Name"
    sheet['K1']="Father's Occupation"
    sheet['L1']="Mother's Occupation"

    file.save('Student_data.xlsx')


#exit
def Exit():
    root.destroy()

#Show image
def showimage():
    global filename
    global img
    filename=filedialog.askopenfilename(initialdir=os.getcwd(),
                                        title="Select image file",
                                        filetypes=(("JPG File","*.jpg"),
                                                   ("PNG File","*.png"),
                                                   ("ALL files","*.txt")))
    img = (Image.open(filename))
    resized_image = img.resize((165,165))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2


#Resigter No
def registration_no():
    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet =  file.active
    row = sheet.max_row

    max_row_value = sheet.cell(row=row,column=1).value
    try :
        Registration.set(int(max_row_value)+1)

    except:
        Registration.set("1")


#Clear
def clear():
    Name.set('')
    DOB.set('')
    Religon.set('')
    Skills.set('')
    F_Name.set('')
    M_Name.set('')
    F_Occupation.set('')
    M_Occupation.set('')
    Class.set("Select Class")

    registration_no()

    saveButton.config(state = 'normal')

    img1 = PhotoImage(file='images/upload photo.png')
    lbl.config(image=img1)
    lbl.image=img1

    img = ""




#gender
def selection():
    global gender
    value=radio.get()
    if value ==1:
        gender="Male"
    else:
        gender="Female"


#Save Function
def save():
    global G1
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    try:
        G1 = gender

    except:
        messagebox.showerror("Error","Select Gender")

    D2 = DOB.get()
    D1 = Date.get()
    rel = Religon.get()
    S1 = Skills.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = F_Occupation.get()
    M1 = M_Occupation.get()

    if N1=="" or C1=="Select Class" or D2=="" or rel=="" or S1=="" or fathername=="" or mothername=="" or F1=="" or M1=="" :
        messagebox.showerror("error","Few Data is missing")
    else:
        file=openpyxl.load_workbook('Student_data.xlsx')
        sheet = file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=R1)
        sheet.cell(column=2,row=sheet.max_row,value=N1)
        sheet.cell(column=3,row=sheet.max_row,value=C1)
        sheet.cell(column=4,row=sheet.max_row,value=G1)
        sheet.cell(column=5,row=sheet.max_row,value=D2)
        sheet.cell(column=6,row=sheet.max_row,value=D1)
        sheet.cell(column=7,row=sheet.max_row,value=rel)
        sheet.cell(column=8,row=sheet.max_row,value=S1)
        sheet.cell(column=9,row=sheet.max_row,value=fathername)
        sheet.cell(column=10,row=sheet.max_row,value=mothername)
        sheet.cell(column=11,row=sheet.max_row,value=F1)
        sheet.cell(column=12,row=sheet.max_row,value=M1)
        file.save(r'Student_data.xlsx')

        try:
            img.save("images/studentImage/"+str(R1)+".jpg")
        except:
            messagebox.showinfo("info","Profile picture is not available!!!")

        messagebox.showinfo("info","Sucessfully data enterd!!")

        clear()
        registration_no()

#search
def search():
    text = Search.get()
    clear()
    saveButton.config(state='disabled')

    file = openpyxl.load_workbook("Student_data.xlsx")
    sheet = file.active
    for row in sheet.rows:
        if row[0].value == int(text):
            name=row[0]
            reg_No_postion=str(name)[14:-1]
            reg_number = str(name)[15:-1]

    try:
        print(str(name))
    except:
        messagebox.showerror("Invaild","Invaild registration number!!")


    x1 = sheet.cell(row=int(reg_number),column=1).value
    x2 = sheet.cell(row=int(reg_number),column=2).value
    x3 = sheet.cell(row=int(reg_number),column=3).value
    x4 = sheet.cell(row=int(reg_number),column=4).value
    x5 = sheet.cell(row=int(reg_number),column=5).value
    x6 = sheet.cell(row=int(reg_number),column=6).value
    x7 = sheet.cell(row=int(reg_number),column=7).value
    x8 = sheet.cell(row=int(reg_number),column=8).value
    x9 = sheet.cell(row=int(reg_number),column=9).value
    x10 = sheet.cell(row=int(reg_number),column=10).value
    x11 = sheet.cell(row=int(reg_number),column=11).value
    x12 = sheet.cell(row=int(reg_number),column=12).value

    Registration.set(x1)
    Name.set(x2)
    Class.set(x3)

    if x4=='Female':
        R2.select()
    else:
        R1.select()

    Date.set(x6)
    DOB.set(x5)
    Religon.set(x7)
    Skills.set(x8)
    F_Name.set(x9)
    M_Name.set(x10)
    F_Occupation.set(x11)
    M_Occupation.set(x12)

    img = (Image.open("images/studentImage/"+str(x1)+".jpg"))
    resized_image  = img.resize((165,165))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2




#top frames
Label(root,text="Email:ijse@gmail.com",width=8,height=2,bg="#f0687c",anchor='e').pack(side=TOP,fill=X)
Label(root,text="STUDENT REGISTRATION",width=8,height=2,bg="#c36464",fg='#fff',font='arial 15 bold').pack(side=TOP,fill=X)

#search box to update
Search = StringVar()
Entry(root,textvariable=Search,width=20,bd=2,font='arial 16').place(x=800,y=50)
imageicon3=PhotoImage(file="images/search.png")
srch=Button(root,text="Search",compound=LEFT,width=100,image=imageicon3,bg='#68ddfa',font='arial 10 bold',command=search)
srch.place(x=1060,y=51)

imageicon4= PhotoImage(file="images/layers.png")
Update_button=Button(root,image=imageicon4,bg="#c36464")
Update_button.place(x=90,y=47)


#Registration and date
Label(root,text="Registration No:",font='arial 10',fg=framebg,bg=background).place(x=28,y=120)
Label(root,text="Date:",font='arial 10',fg=framebg,bg=background).place(x=490,y=120)

Registration=IntVar()
Date = StringVar()

reg_entry = Entry(root,textvariable=Registration,width=15,font='arial 12')
reg_entry.place(x=145,y=120)

registration_no()


#Registration date()
today=date.today()
d1= today.strftime("%d/%m/%y")
date_entry = Entry(root,textvariable=Date,width=15,font='arial 12')
date_entry.place(x=540,y=120)
Date.set(d1)

#Student detail
obj=LabelFrame(root,text="Student's Deatil",font=18,bd=2,width=880,bg=framebg,fg=framefg,height=210,relief=GROOVE)
obj.place(x=20,y=170)

Label(obj,text="Full name:",font='arial 12',bg=framebg,fg=framefg).place(x=20,y=35)
Label(obj,text="Date of Birth:",font='arial 12',bg=framebg,fg=framefg).place(x=20,y=85)
Label(obj,text="Gender:",font='arial 12',bg=framebg,fg=framefg).place(x=20,y=135)

Label(obj,text="Class:",font='arial 12',bg=framebg,fg=framefg).place(x=480,y=35)
Label(obj,text="Religon:",font='arial 12',bg=framebg,fg=framefg).place(x=480,y=85)
Label(obj,text="Skills:",font='arial 12',bg=framebg,fg=framefg).place(x=480,y=135)

Name = StringVar()
name_entry = Entry(obj,textvariable=Name,width=25,font='arial 8')
name_entry.place(x=140,y=35)

DOB = StringVar()
dob_entry = Entry(obj,textvariable=DOB,width=25,font='arial 8')
dob_entry.place(x=140,y=85)

radio = IntVar()
R1= Radiobutton(obj,text="Male",variable=radio,value=1,bg=framebg,fg=framefg,command=selection)
R1.place(x=130,y=135)
R2= Radiobutton(obj,text="Female",variable=radio,value=2,bg=framebg,fg=framefg,command=selection)
R2.place(x=200,y=135)

Class = Combobox(obj,values=['1','2','3','4','5','6','7','8','9','10','11','12','13'],font='Roboto 8',width=23,state='r')
Class.place(x=600,y=35)
Class.set("Select Class")

Religon = StringVar()
religon_entry = Entry(obj,textvariable=Religon,width=25,font='arial 8')
religon_entry.place(x=600,y=85)

Skills = StringVar()
skills_entry = Entry(obj,textvariable=Skills,width=25,font='arial 8')
skills_entry.place(x=600,y=135)

#Parents detail
obj2=LabelFrame(root,text="Parents' Deatil",font=18,bd=2,width=880,bg=framebg,fg=framefg,height=180,relief=GROOVE)
obj2.place(x=20,y=390)

Label(obj2,text="Father's name:",font='arial 12',bg=framebg,fg=framefg).place(x=20,y=35)
Label(obj2,text="Occupation:",font='arial 12',bg=framebg,fg=framefg).place(x=20,y=85)

Label(obj2,text="Mother's name:",font='arial 12',bg=framebg,fg=framefg).place(x=480,y=35)
Label(obj2,text="Occupation:",font='arial 12',bg=framebg,fg=framefg).place(x=480,y=85)

F_Name=StringVar()
f_entry=Entry(obj2,textvariable=F_Name,width=25,font='arial 8')
f_entry.place(x=140,y=35)

F_Occupation=StringVar()
FO_entry=Entry(obj2,textvariable=F_Occupation,width=25,font='arial 8')
FO_entry.place(x=140,y=85)

M_Name=StringVar()
m_entry=Entry(obj2,textvariable=M_Name,width=25,font='arial 8')
m_entry.place(x=600,y=35)

M_Occupation=StringVar()
MO_entry=Entry(obj2,textvariable=M_Occupation,width=25,font='arial 8')
MO_entry.place(x=600,y=85)

#image
f=Frame(root,bd=3,bg="Black",width=165,height=165,relief=GROOVE)
f.place(x=970,y=120)

img=PhotoImage(file="images/upload photo.png")
lbl=Label(f,bg="black",image=img)
lbl.place(x=0,y=0)


#button
Button(root,text="Upload",width=19,height=2,font='arial 12 bold',bg='lightblue',command=showimage).place(x=970,y=305)

saveButton = Button(root,text="Save",width=19,height=2,font='arial 12 bold',bg='lightgreen',command=save)
saveButton.place(x=970,y=375)

Button(root,text="Reset",width=19,height=2,font='arial 12 bold',bg='lightpink',command=clear).place(x=970,y=445)

Button(root,text="Exit",width=19,height=2,font='arial 12 bold',bg='grey',command=Exit).place(x=970,y=515)



root.mainloop()











